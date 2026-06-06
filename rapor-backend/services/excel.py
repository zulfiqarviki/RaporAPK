from datetime import datetime
from io import BytesIO

from fastapi import HTTPException, status, UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError

from models.grade_component import GradeComponent
from models.score import Score
from models.student import Student
from models.user import User
from services.grade_table import get_grade_table_by_id, GradeTable
from schemas.excel import ExcelImportOut

def export_grade_table_to_excel(
    grade_table_id: int,
    current_user: User,
    db: Session,
) -> BytesIO:
    grade_table = get_grade_table_by_id(
        grade_table_id=grade_table_id,
        current_user=current_user,
        db=db,
    )

    components = (
        db.query(GradeComponent)
        .filter(GradeComponent.grade_table_id == grade_table.id)
        .order_by(GradeComponent.order_index.asc(), GradeComponent.id.asc())
        .all()
    )

    students = (
        db.query(Student)
        .filter(Student.grade_table_id == grade_table.id)
        .order_by(Student.id.asc())
        .all()
    )

    scores = (
        db.query(Score)
        .join(Student, Score.student_id == Student.id)
        .filter(Student.grade_table_id == grade_table.id)
        .all()
    )

    score_map = {
        (score.student_id, score.component_id): score.score
        for score in scores
    }

    workbook = Workbook()

    metadata_sheet = workbook.active
    metadata_sheet.title = "Metadata"

    _write_metadata_sheet(
        sheet=metadata_sheet,
        grade_table=grade_table,
    )

    components_sheet = workbook.create_sheet("Components")
    _write_components_sheet(
        sheet=components_sheet,
        components=components,
    )

    scores_sheet = workbook.create_sheet("Scores")
    _write_scores_sheet(
        sheet=scores_sheet,
        students=students,
        components=components,
        score_map=score_map,
    )

    results_sheet = workbook.create_sheet("Results")
    _write_results_sheet(
        sheet=results_sheet,
        students=students,
        components=components,
        score_map=score_map,
    )

    for sheet in workbook.worksheets:
        _style_header_row(sheet)
        _auto_fit_columns(sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


def _write_metadata_sheet(sheet, grade_table) -> None:
    rows = [
        ("Grade Table ID", grade_table.id),
        ("Subject Name", grade_table.subject_name),
        ("Description", grade_table.description or ""),
        ("Teacher ID", grade_table.teacher_id),
        ("Teacher NIP", grade_table.teacher.nip),
        ("Teacher Name", grade_table.teacher.name),
        ("Exported At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    sheet.append(["Field", "Value"])

    for row in rows:
        sheet.append(list(row))


def _write_components_sheet(
    sheet,
    components: list[GradeComponent],
) -> None:
    sheet.append(
        [
            "Component ID",
            "Component Name",
            "Weight",
            "Max Score",
            "Order Index",
        ]
    )

    for component in components:
        sheet.append(
            [
                component.id,
                component.name,
                component.weight,
                component.max_score,
                component.order_index,
            ]
        )


def _write_scores_sheet(
    sheet,
    students: list[Student],
    components: list[GradeComponent],
    score_map: dict[tuple[int, int], float],
) -> None:
    header = [
        "Student ID",
        "Student Number",
        "Student Name",
    ]

    for component in components:
        header.append(component.name)

    sheet.append(header)

    for student in students:
        row = [
            student.id,
            student.student_number or "",
            student.name,
        ]

        for component in components:
            score_value = score_map.get((student.id, component.id))

            # Raw score is left blank if missing.
            # This preserves the difference between actual 0 and missing score.
            row.append(score_value if score_value is not None else "")

        sheet.append(row)


def _write_results_sheet(
    sheet,
    students: list[Student],
    components: list[GradeComponent],
    score_map: dict[tuple[int, int], float],
) -> None:
    sheet.append(
        [
            "Student ID",
            "Student Number",
            "Student Name",
            "Final Grade",
            "Is Complete",
            "Missing Components",
        ]
    )

    total_weight = sum(component.weight for component in components)

    if components and total_weight <= 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Total component weight must be greater than 0",
        )

    for student in students:
        final_grade, is_complete, missing_components = _calculate_final_grade(
            student=student,
            components=components,
            score_map=score_map,
            total_weight=total_weight,
        )

        sheet.append(
            [
                student.id,
                student.student_number or "",
                student.name,
                final_grade,
                "Yes" if is_complete else "No",
                ", ".join(missing_components),
            ]
        )


def _calculate_final_grade(
    student: Student,
    components: list[GradeComponent],
    score_map: dict[tuple[int, int], float],
    total_weight: float,
) -> tuple[float | None, bool, list[str]]:
    if not components:
        return None, False, []

    weighted_total = 0.0
    missing_components: list[str] = []

    for component in components:
        score_value = score_map.get((student.id, component.id))

        if score_value is None:
            effective_score = 0.0
            missing_components.append(component.name)
        else:
            effective_score = score_value

        weighted_total += effective_score * component.weight

    final_grade = round(weighted_total / total_weight, 2)
    is_complete = len(missing_components) == 0

    return final_grade, is_complete, missing_components


def _style_header_row(sheet) -> None:
    if sheet.max_row < 1:
        return

    for cell in sheet[1]:
        cell.font = Font(bold=True)


def _auto_fit_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is None:
                continue

            value_length = len(str(cell.value))

            if value_length > max_length:
                max_length = value_length

        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width


def import_grade_table_from_excel(
    file: UploadFile,
    current_user: User,
    db: Session,
    teacher_id: int | None = None,
) -> ExcelImportOut:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are supported",
        )

    owner_id = _resolve_import_owner_id(
        current_user=current_user,
        teacher_id=teacher_id,
        db=db,
    )

    try:
        file.file.seek(0)
        workbook = load_workbook(file.file, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Excel file",
        )

    # Validation buffer.
    # Database is not changed before this returns successfully.
    parsed_data = _parse_and_validate_import_workbook(workbook)

    try:
        grade_table = GradeTable(
            subject_name=parsed_data["subject_name"],
            description=parsed_data["description"],
            teacher_id=owner_id,
        )

        db.add(grade_table)
        db.flush()

        component_map: dict[str, GradeComponent] = {}

        for component_data in parsed_data["components"]:
            component = GradeComponent(
                name=component_data["name"],
                weight=component_data["weight"],
                max_score=100,
                order_index=component_data["order_index"],
                grade_table_id=grade_table.id,
            )

            db.add(component)
            db.flush()

            component_map[component.name] = component

        imported_students = 0
        imported_scores = 0

        for student_data in parsed_data["students"]:
            student = Student(
                name=student_data["student_name"],
                student_number=student_data["student_number"],
                grade_table_id=grade_table.id,
            )

            db.add(student)
            db.flush()

            imported_students += 1

            for component_name, score_value in student_data["scores"].items():
                if score_value is None:
                    continue

                component = component_map[component_name]

                score = Score(
                    student_id=student.id,
                    component_id=component.id,
                    score=score_value,
                )

                db.add(score)
                imported_scores += 1

        db.commit()
        db.refresh(grade_table)

    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Imported data violates database constraints",
        )

    except HTTPException:
        db.rollback()
        raise

    except Exception:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to import Excel file",
        )

    return ExcelImportOut(
        grade_table_id=grade_table.id,
        subject_name=grade_table.subject_name,
        teacher_id=grade_table.teacher_id,
        imported_components=len(parsed_data["components"]),
        imported_students=imported_students,
        imported_scores=imported_scores,
    )


def _resolve_import_owner_id(
    current_user: User,
    teacher_id: int | None,
    db: Session,
) -> int:
    if current_user.role == "admin":
        if teacher_id is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="teacher_id is required when admin imports a grade table",
            )

        teacher = (
            db.query(User)
            .filter(User.id == teacher_id, User.role == "teacher")
            .first()
        )

        if teacher is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Teacher not found",
            )

        return teacher.id

    if teacher_id is not None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Teacher cannot import grade table for another teacher",
        )

    return current_user.id


def _parse_and_validate_import_workbook(workbook) -> dict:
    _validate_required_import_sheets(workbook)

    metadata = _read_import_metadata_sheet(workbook["Metadata"])
    components = _read_import_components_sheet(workbook["Components"])
    students = _read_import_scores_sheet(
        sheet=workbook["Scores"],
        component_names=[component["name"] for component in components],
    )

    subject_name = metadata.get("Subject Name")

    if subject_name is None or subject_name.strip() == "":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Metadata sheet must contain non-empty Subject Name",
        )

    description = metadata.get("Description")

    return {
        "subject_name": subject_name.strip(),
        "description": description.strip() if description else None,
        "components": components,
        "students": students,
    }


def _validate_required_import_sheets(workbook) -> None:
    required_sheets = {"Metadata", "Components", "Scores"}
    existing_sheets = set(workbook.sheetnames)

    missing_sheets = required_sheets - existing_sheets

    if missing_sheets:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required sheets: {', '.join(sorted(missing_sheets))}",
        )


def _read_import_metadata_sheet(sheet) -> dict[str, str]:
    header = _get_import_header_map(sheet)

    _validate_exact_headers(
        sheet_name="Metadata",
        actual_headers=set(header.keys()),
        required_headers={"Field", "Value"},
        optional_headers=set(),
    )

    metadata: dict[str, str] = {}
    allowed_fields = {
        "Subject Name",
        "Description",
        "Grade Table ID",
        "Teacher ID",
        "Teacher NIP",
        "Teacher Name",
        "Exported At",
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        field = row[header["Field"]]
        value = row[header["Value"]]

        if field is None or str(field).strip() == "":
            continue

        field_name = str(field).strip()

        if field_name not in allowed_fields:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Metadata row {row_index}: Unknown metadata field '{field_name}'",
            )

        if field_name in metadata:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Metadata sheet contains duplicate field: {field_name}",
            )

        metadata[field_name] = "" if value is None else str(value).strip()

    if "Subject Name" not in metadata:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Metadata sheet must contain Subject Name",
        )

    return metadata


def _read_import_components_sheet(sheet) -> list[dict]:
    header = _get_import_header_map(sheet)

    _validate_exact_headers(
        sheet_name="Components",
        actual_headers=set(header.keys()),
        required_headers={"Component Name", "Weight", "Max Score", "Order Index"},
        optional_headers={"Component ID"},
    )

    component_rows: list[dict] = []
    seen_component_names: set[str] = set()

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_component_name = row[header["Component Name"]]

        if raw_component_name is None or str(raw_component_name).strip() == "":
            continue

        component_name = str(raw_component_name).strip()

        if component_name in seen_component_names:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Components sheet contains duplicate component name: {component_name}",
            )

        weight = _parse_import_float(
            value=row[header["Weight"]],
            sheet_name="Components",
            row_index=row_index,
            field_name="Weight",
        )

        max_score = _parse_import_float(
            value=row[header["Max Score"]],
            sheet_name="Components",
            row_index=row_index,
            field_name="Max Score",
        )

        order_index = _parse_import_int(
            value=row[header["Order Index"]],
            sheet_name="Components",
            row_index=row_index,
            field_name="Order Index",
        )

        if weight <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Components row {row_index}: Weight must be greater than 0",
            )

        if max_score != 100:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Components row {row_index}: Max Score must be 100",
            )

        seen_component_names.add(component_name)

        component_rows.append(
            {
                "name": component_name,
                "weight": weight,
                "order_index": order_index,
            }
        )

    if not component_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Components sheet must contain at least one component",
        )

    return component_rows


def _read_import_scores_sheet(
    sheet,
    component_names: list[str],
) -> list[dict]:
    header = _get_import_header_map(sheet)

    required_headers = {"Student Number", "Student Name"}
    optional_headers = {"Student ID"}
    expected_component_headers = set(component_names)

    _validate_exact_headers(
        sheet_name="Scores",
        actual_headers=set(header.keys()),
        required_headers=required_headers | expected_component_headers,
        optional_headers=optional_headers,
    )

    score_component_headers = {
        header_name
        for header_name in header
        if header_name not in {"Student ID", "Student Number", "Student Name"}
    }

    missing_component_columns = expected_component_headers - score_component_headers
    extra_component_columns = score_component_headers - expected_component_headers

    if missing_component_columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Scores sheet missing component columns: {', '.join(sorted(missing_component_columns))}",
        )

    if extra_component_columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Scores sheet has unknown component columns: {', '.join(sorted(extra_component_columns))}",
        )

    score_rows: list[dict] = []
    seen_student_numbers: set[str] = set()

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_student_name = row[header["Student Name"]]

        if raw_student_name is None or str(raw_student_name).strip() == "":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Scores row {row_index}: Student Name is required",
            )

        student_name = str(raw_student_name).strip()

        raw_student_number = row[header["Student Number"]]
        student_number = (
            None
            if raw_student_number is None or str(raw_student_number).strip() == ""
            else str(raw_student_number).strip()
        )

        if student_number is not None:
            if student_number in seen_student_numbers:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Scores row {row_index}: Duplicate student number '{student_number}'",
                )

            seen_student_numbers.add(student_number)

        scores: dict[str, float | None] = {}

        for component_name in component_names:
            raw_score = row[header[component_name]]

            if raw_score is None or raw_score == "":
                scores[component_name] = None
                continue

            parsed_score = _parse_import_float(
                value=raw_score,
                sheet_name="Scores",
                row_index=row_index,
                field_name=component_name,
            )

            if parsed_score < 0 or parsed_score > 100:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Scores row {row_index}: {component_name} must be between 0 and 100",
                )

            scores[component_name] = parsed_score

        score_rows.append(
            {
                "student_number": student_number,
                "student_name": student_name,
                "scores": scores,
            }
        )

    if not score_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Scores sheet must contain at least one student",
        )

    return score_rows


def _get_import_header_map(sheet) -> dict[str, int]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)

    if header_row is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{sheet.title} sheet is empty",
        )

    header_map: dict[str, int] = {}

    for index, value in enumerate(header_row):
        if value is None or str(value).strip() == "":
            continue

        header_name = str(value).strip()

        if header_name in header_map:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"{sheet.title} sheet contains duplicate header: {header_name}",
            )

        header_map[header_name] = index

    if not header_map:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{sheet.title} sheet must contain headers",
        )

    return header_map


def _validate_exact_headers(
    sheet_name: str,
    actual_headers: set[str],
    required_headers: set[str],
    optional_headers: set[str],
) -> None:
    missing_headers = required_headers - actual_headers

    if missing_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{sheet_name} sheet missing required headers: {', '.join(sorted(missing_headers))}",
        )

    allowed_headers = required_headers | optional_headers
    unknown_headers = actual_headers - allowed_headers

    if unknown_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{sheet_name} sheet has unknown headers: {', '.join(sorted(unknown_headers))}",
        )


def _parse_import_float(
    value,
    sheet_name: str,
    row_index: int,
    field_name: str,
) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{sheet_name} row {row_index}: {field_name} must be numeric",
        )


def _parse_import_int(
    value,
    sheet_name: str,
    row_index: int,
    field_name: str,
) -> int:
    try:
        converted_value = int(value)
    except (TypeError, ValueError):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{sheet_name} row {row_index}: {field_name} must be an integer",
        )

    return converted_value